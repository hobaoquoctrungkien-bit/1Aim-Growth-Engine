from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile
from xml.sax.saxutils import escape


def money_text(value, currency="USD"):
    try:
        return f"{currency} {float(value or 0):,.2f}"
    except (TypeError, ValueError):
        return f"{currency} 0.00"


def quote_title(quotation):
    quote_no = quotation.get("quote_no") or "Draft"
    version = quotation.get("version") or 1
    return f"Quotation {quote_no} v{version}"


def build_quote_rows(quotation):
    rows = []
    for item in quotation.get("items", []):
        rows.append(
            {
                "Line": item.get("line_no"),
                "Description": item.get("description"),
                "Basis": item.get("basis"),
                "Qty": item.get("quantity"),
                "Unit Price": item.get("unit_price"),
                "Currency": item.get("currency") or quotation.get("currency") or "USD",
                "Amount": item.get("amount"),
                "Vendor": item.get("vendor_name"),
                "Notes": item.get("notes"),
            }
        )
    return rows


def build_quotation_excel(quotation):
    try:
        return build_quotation_excel_with_openpyxl(quotation)
    except ModuleNotFoundError:
        return build_quotation_xlsx_minimal(quotation)


def build_quotation_excel_with_openpyxl(quotation):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quotation"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(color="FFFFFF", bold=True, size=16)
    bold_font = Font(bold=True)

    sheet.merge_cells("A1:I1")
    sheet["A1"] = quote_title(quotation)
    sheet["A1"].font = title_font
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center")

    metadata = [
        ("Customer", quotation.get("customer_name")),
        ("Contact", quotation.get("contact_name")),
        ("Opportunity", quotation.get("opportunity_name")),
        ("Trade Lane", quotation.get("trade_lane")),
        ("Service", quotation.get("service_type")),
        ("Quote Date", quotation.get("quote_date")),
        ("Valid Until", quotation.get("valid_until")),
        ("Status", quotation.get("status")),
        ("Payment Terms", quotation.get("payment_terms")),
    ]
    row_index = 3
    for label, value in metadata:
        sheet.cell(row=row_index, column=1, value=label).font = bold_font
        sheet.cell(row=row_index, column=2, value=value or "-")
        row_index += 1

    row_index += 1
    headers = ["Line", "Description", "Basis", "Qty", "Unit Price", "Currency", "Amount", "Vendor", "Notes"]
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=col_index, value=header)
        cell.font = bold_font
        cell.fill = header_fill
    row_index += 1

    for item in build_quote_rows(quotation):
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index, column=col_index, value=item.get(header))
        row_index += 1

    row_index += 1
    sheet.cell(row=row_index, column=6, value="Total").font = bold_font
    sheet.cell(row=row_index, column=7, value=quotation.get("sell_amount") or quotation.get("totals", {}).get("sell_amount") or 0).font = bold_font
    sheet.cell(row=row_index, column=8, value=quotation.get("currency") or "USD").font = bold_font

    widths = [12, 34, 16, 10, 14, 12, 14, 22, 34]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def xlsx_cell_ref(row, column):
    letters = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row}"


def xlsx_inline_cell(row, column, value, style_id=0):
    cell_ref = xlsx_cell_ref(row, column)
    style = f' s="{style_id}"' if style_id else ""
    if value in [None, ""]:
        value = "-"
    if isinstance(value, (int, float)):
        return f'<c r="{cell_ref}"{style}><v>{value}</v></c>'
    return f'<c r="{cell_ref}" t="inlineStr"{style}><is><t>{escape(str(value))}</t></is></c>'


def build_quotation_xlsx_minimal(quotation):
    rows = []
    rows.append([(quote_title(quotation), 1)])
    rows.append([])
    metadata = [
        ("Customer", quotation.get("customer_name")),
        ("Contact", quotation.get("contact_name")),
        ("Opportunity", quotation.get("opportunity_name")),
        ("Trade Lane", quotation.get("trade_lane")),
        ("Service", quotation.get("service_type")),
        ("Quote Date", quotation.get("quote_date")),
        ("Valid Until", quotation.get("valid_until")),
        ("Status", quotation.get("status")),
        ("Payment Terms", quotation.get("payment_terms")),
    ]
    for label, value in metadata:
        rows.append([(label, 2), (value or "-", 0)])
    rows.append([])
    headers = ["Line", "Description", "Basis", "Qty", "Unit Price", "Currency", "Amount", "Vendor", "Notes"]
    rows.append([(header, 2) for header in headers])
    for item in build_quote_rows(quotation):
        rows.append([(item.get(header), 0) for header in headers])
    rows.append([])
    rows.append([
        ("", 0),
        ("", 0),
        ("", 0),
        ("", 0),
        ("", 0),
        ("Total", 2),
        (quotation.get("sell_amount") or quotation.get("totals", {}).get("sell_amount") or 0, 2),
        (quotation.get("currency") or "USD", 2),
    ])

    sheet_rows = []
    for row_index, row in enumerate(rows, start=1):
        cells = []
        for column_index, cell in enumerate(row, start=1):
            if isinstance(cell, tuple):
                value, style_id = cell
            else:
                value, style_id = cell, 0
            cells.append(xlsx_inline_cell(row_index, column_index, value, style_id))
        sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')

    sheet_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <cols>
    <col min="1" max="1" width="12" customWidth="1"/>
    <col min="2" max="2" width="34" customWidth="1"/>
    <col min="3" max="3" width="16" customWidth="1"/>
    <col min="4" max="4" width="10" customWidth="1"/>
    <col min="5" max="5" width="14" customWidth="1"/>
    <col min="6" max="6" width="12" customWidth="1"/>
    <col min="7" max="7" width="14" customWidth="1"/>
    <col min="8" max="8" width="22" customWidth="1"/>
    <col min="9" max="9" width="34" customWidth="1"/>
  </cols>
  <sheetData>{''.join(sheet_rows)}</sheetData>
</worksheet>"""
    workbook_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="Quotation" sheetId="1" r:id="rId1"/></sheets>
</workbook>"""
    workbook_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""
    styles_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="3"><font/><font><b/><sz val="16"/><color rgb="FFFFFFFF"/></font><font><b/></font></fonts>
  <fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FF1F4E78"/></patternFill></fill></fills>
  <borders count="1"><border/></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="3"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/><xf numFmtId="0" fontId="1" fillId="2" borderId="0" applyFont="1" applyFill="1"/><xf numFmtId="0" fontId="2" fillId="0" borderId="0" applyFont="1"/></cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>"""
    root_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>"""

    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", root_rels)
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        archive.writestr("xl/styles.xml", styles_xml)
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return output.getvalue()


def pdf_escape(text):
    return str(text or "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def build_pdf_stream(lines):
    y = 780
    commands = ["BT", "/F1 11 Tf"]
    for line in lines:
        commands.append(f"1 0 0 1 50 {y} Tm")
        commands.append(f"({pdf_escape(line)}) Tj")
        y -= 16
    commands.append("ET")
    return "\n".join(commands).encode("latin-1", errors="replace")


def build_quotation_pdf(quotation):
    currency = quotation.get("currency") or "USD"
    total = quotation.get("sell_amount") or quotation.get("totals", {}).get("sell_amount") or 0
    lines = [
        quote_title(quotation),
        f"Customer: {quotation.get('customer_name') or '-'}",
        f"Contact: {quotation.get('contact_name') or '-'}",
        f"Trade Lane: {quotation.get('trade_lane') or '-'}",
        f"Service: {quotation.get('service_type') or '-'}",
        f"Quote Date: {quotation.get('quote_date') or '-'}",
        f"Valid Until: {quotation.get('valid_until') or '-'}",
        f"Status: {quotation.get('status') or '-'}",
        "",
        "Charges",
    ]
    for item in quotation.get("items", []):
        lines.append(
            f"{item.get('line_no')}. {item.get('description')} - "
            f"{money_text(item.get('amount'), item.get('currency') or currency)}"
        )
    lines.extend(
        [
            "",
            f"Total: {money_text(total, currency)}",
            f"Payment Terms: {quotation.get('payment_terms') or '-'}",
            f"Notes: {quotation.get('notes') or '-'}",
        ]
    )

    stream = build_pdf_stream(lines[:42])
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    pdf = BytesIO()
    pdf.write(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(pdf.tell())
        pdf.write(f"{index} 0 obj\n".encode("ascii"))
        pdf.write(obj)
        pdf.write(b"\nendobj\n")
    xref_at = pdf.tell()
    pdf.write(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.write(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.write(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.write(
        (
            f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_at}\n%%EOF\n"
        ).encode("ascii")
    )
    return pdf.getvalue()
